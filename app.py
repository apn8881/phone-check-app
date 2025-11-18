import streamlit as st
import pandas as pd
import sqlite3
import io
from datetime import datetime
import openpyxl
import os
import shutil
import tempfile
import gdown
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive

# ตั้งค่าหน้า
st.set_page_config(
    page_title="โปรแกรมเช็คเบอร์โทรซ้ำ - Google Drive",
    page_icon="📱",
    layout="wide"
)

# รหัสผ่าน
PASSWORD = "23669"

# Google Drive Setup
def setup_google_drive():
    """ตั้งค่า Google Drive"""
    try:
        gauth = GoogleAuth()
        
        # ลองใช้ credentials ที่มีอยู่
        gauth.LoadCredentialsFile("google_drive_credentials.json")
        
        if gauth.credentials is None:
            # ถ้าไม่มี credentials ให้ผู้ใช้กำหนดค่า
            st.warning("🔐 ต้องการตั้งค่า Google Drive")
            st.info("""
            **ขั้นตอนการตั้งค่า:**
            1. ไปที่ [Google Cloud Console](https://console.cloud.google.com/)
            2. สร้างโปรเจคใหม่
            3. เปิดใช้งาน Google Drive API
            4. สร้าง credentials (OAuth 2.0 Client ID)
            5. ดาวน์โหลดไฟล์ client_secrets.json
            6. อัพโหลดไฟล์ client_secrets.json ด้านล่าง
            """)
            
            uploaded_secrets = st.file_uploader("อัพโหลด client_secrets.json", type=['json'])
            if uploaded_secrets:
                with open("client_secrets.json", "wb") as f:
                    f.write(uploaded_secrets.getvalue())
                
                gauth = GoogleAuth()
                gauth.LocalWebserverAuth()
                gauth.SaveCredentialsFile("google_drive_credentials.json")
                st.success("✅ ตั้งค่า Google Drive สำเร็จ!")
        
        drive = GoogleDrive(gauth)
        return drive
    except Exception as e:
        st.error(f"❌ ตั้งค่า Google Drive ไม่สำเร็จ: {str(e)}")
        return None

# ฟังก์ชันจัดการ Google Drive
def get_database_from_drive(drive, file_id=None):
    """ดาวน์โหลดฐานข้อมูลจาก Google Drive"""
    try:
        # ค้นหาไฟล์ใน Google Drive
        if file_id:
            file = drive.CreateFile({'id': file_id})
        else:
            file_list = drive.ListFile({'q': "title='phone_database.db' and trashed=false"}).GetList()
            if file_list:
                file = file_list[0]
            else:
                return None
        
        # ดาวน์โหลดไฟล์
        local_path = "phone_database.db"
        file.GetContentFile(local_path)
        return local_path
    except Exception as e:
        st.error(f"❌ ดาวน์โหลดจาก Google Drive ไม่สำเร็จ: {str(e)}")
        return None

def upload_database_to_drive(drive, local_path, file_id=None):
    """อัพโหลดฐานข้อมูลไปยัง Google Drive"""
    try:
        if file_id:
            # อัพเดทไฟล์ที่มีอยู่
            file = drive.CreateFile({'id': file_id})
        else:
            # สร้างไฟล์ใหม่
            file = drive.CreateFile({'title': 'phone_database.db'})
        
        file.SetContentFile(local_path)
        file.Upload()
        return file['id']
    except Exception as e:
        st.error(f"❌ อัพโหลดไป Google Drive ไม่สำเร็จ: {str(e)}")
        return None

def sync_with_drive():
    """ซิงค์ข้อมูลกับ Google Drive"""
    drive = setup_google_drive()
    if not drive:
        return None, None
    
    local_db_path = "phone_database.db"
    drive_file_id = None
    
    # ค้นหาไฟล์ใน Drive
    try:
        file_list = drive.ListFile({'q': "title='phone_database.db' and trashed=false"}).GetList()
        if file_list:
            drive_file_id = file_list[0]['id']
            
            # ตรวจสอบว่าไฟล์ไหนใหม่กว่า
            local_time = os.path.getmtime(local_db_path) if os.path.exists(local_db_path) else 0
            drive_time = file_list[0]['modifiedDate']
            drive_time = datetime.strptime(drive_time, '%Y-%m-%dT%H:%M:%S.%fZ').timestamp()
            
            if drive_time > local_time:
                # ดาวน์โหลดจาก Drive
                get_database_from_drive(drive, drive_file_id)
                st.sidebar.success("✅ โหลดข้อมูลจาก Google Drive สำเร็จ")
            else:
                # อัพโหลดไป Drive
                upload_database_to_drive(drive, local_db_path, drive_file_id)
                st.sidebar.success("✅ บันทึกข้อมูลไป Google Drive สำเร็จ")
        else:
            # สร้างไฟล์ใหม่ใน Drive
            if os.path.exists(local_db_path):
                drive_file_id = upload_database_to_drive(drive, local_db_path)
                st.sidebar.success("✅ สร้างไฟล์ใหม่ใน Google Drive สำเร็จ")
    except Exception as e:
        st.sidebar.error(f"❌ ซิงค์ข้อมูลไม่สำเร็จ: {str(e)}")
    
    return drive, drive_file_id

# ฟังก์ชันจัดการฐานข้อมูล
def init_database():
    """สร้างฐานข้อมูล SQLite"""
    conn = sqlite3.connect("phone_database.db", timeout=30)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS old_phones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone_number TEXT,
            last_9_digits TEXT UNIQUE,
            source_file TEXT,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_last_9_digits ON old_phones(last_9_digits)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_created_date ON old_phones(created_date)')
    
    conn.commit()
    conn.close()

def extract_last_9_digits(phone):
    """ดึงตัวเลข 9 ตัวท้ายจากเบอร์โทร"""
    if pd.isna(phone) or phone == '' or phone is None:
        return ""
    
    phone_str = str(phone).strip()
    digits_only = ''.join([c for c in phone_str if c.isdigit()])
    
    return digits_only[-9:] if len(digits_only) >= 9 else digits_only

def get_all_last_9_digits():
    """ดึงตัวเลข 9 ตัวท้ายทั้งหมดจากฐานข้อมูล"""
    conn = sqlite3.connect("phone_database.db", timeout=30)
    cursor = conn.cursor()
    
    cursor.execute("SELECT last_9_digits FROM old_phones WHERE LENGTH(last_9_digits) = 9")
    
    results = set()
    batch_size = 100000
    while True:
        batch = cursor.fetchmany(batch_size)
        if not batch:
            break
        results.update(result[0] for result in batch)
    
    conn.close()
    return results

def get_database_stats():
    """ดึงสถิติจากฐานข้อมูล"""
    conn = sqlite3.connect("phone_database.db", timeout=30)
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM old_phones")
    total_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM old_phones WHERE LENGTH(last_9_digits) = 9")
    valid_count = cursor.fetchone()[0]
    
    conn.close()
    return total_count, valid_count

def save_phones_to_database(phone_numbers, source_file=""):
    """บันทึกเบอร์โทรลงฐานข้อมูล"""
    conn = sqlite3.connect("phone_database.db", timeout=30)
    
    new_records_count = 0
    for phone in phone_numbers:
        last_9 = extract_last_9_digits(phone)
        if len(last_9) == 9:
            try:
                cursor = conn.execute(
                    "INSERT OR IGNORE INTO old_phones (phone_number, last_9_digits, source_file) VALUES (?, ?, ?)",
                    (str(phone), last_9, source_file)
                )
                if cursor.rowcount > 0:
                    new_records_count += 1
            except:
                continue
    
    conn.commit()
    conn.close()
    
    # ซิงค์กับ Google Drive
    if new_records_count > 0:
        sync_with_drive()
    
    return new_records_count

def save_phones_as_excel(df):
    """บันทึก DataFrame เป็น Excel"""
    output = io.BytesIO()
    
    if df.empty or len(df) == 0:
        empty_df = pd.DataFrame(columns=df.columns)
        empty_df.loc[0] = ['ไม่มีข้อมูล'] + [''] * (len(df.columns) - 1)
        df = empty_df
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "เบอร์โทร"
    
    # เขียนหัวข้อ
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # เขียนข้อมูล
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_idx == 1 and pd.notna(value) and value != '':
                phone_str = str(value).strip()
                if phone_str and phone_str != 'ไม่มีข้อมูล':
                    cell.value = phone_str
                    cell.number_format = '@'
                else:
                    cell.value = phone_str
            else:
                if pd.notna(value):
                    cell.value = value
                else:
                    cell.value = ''
    
    column_widths = [20, 15, 20, 15]
    for col_idx, width in enumerate(column_widths[:len(df.columns)], 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    
    wb.save(output)
    output.seek(0)
    return output

# เริ่มต้นแอป
def main():
    st.title("📱 โปรแกรมเช็คเบอร์โทรซ้ำ - Google Drive")
    st.markdown("อัพโหลดไฟล์ Excel เพื่อตรวจสอบเบอร์โทรซ้ำโดยใช้**ตัวเลข 9 ตัวท้าย**")
    
    # ซิงค์กับ Google Drive
    if st.sidebar.button("🔄 ซิงค์กับ Google Drive"):
        with st.spinner("กำลังซิงค์ข้อมูล..."):
            sync_with_drive()
    
    # แสดงสถานะ
    if os.path.exists("phone_database.db"):
        file_size = os.path.getsize("phone_database.db")
        file_time = datetime.fromtimestamp(os.path.getmtime("phone_database.db"))
        st.sidebar.markdown(f"**📁 ขนาดไฟล์:** {file_size:,} bytes")
        st.sidebar.markdown(f"**🕒 อัพเดตล่าสุด:** {file_time.strftime('%Y-%m-%d %H:%M')}")
    
    # สถิติ
    total_count, valid_count = get_database_stats()
    st.sidebar.markdown("---")
    st.sidebar.markdown("**📊 สถิติ:**")
    st.sidebar.markdown(f"เบอร์โทรทั้งหมด: **{total_count:,}**")
    st.sidebar.markdown(f"เบอร์ที่ตรวจสอบได้: **{valid_count:,}**")
    
    # ส่วนหลัก
    st.markdown("---")
    
    # อัพโหลดไฟล์ Excel
    uploaded_file = st.file_uploader(
        "**เลือกไฟล์ Excel**", 
        type=['xlsx', 'xls'],
        help="ไฟล์ Excel ต้องมีคอลัมน์แรกเป็นเบอร์โทร"
    )
    
    if uploaded_file is not None:
        col1, col2 = st.columns(2)
        
        with col1:
            save_to_db = st.checkbox(
                "💾 บันทึกเบอร์จากไฟล์นี้ลงฐานข้อมูล", 
                value=True,
                help="บันทึกเบอร์โทรจากไฟล์นี้เพื่อใช้ตรวจสอบซ้ำในครั้งต่อไป"
            )
        
        with col2:
            if st.button("🚀 เริ่มตรวจสอบเบอร์โทรซ้ำ", type="primary", use_container_width=True):
                with st.spinner('กำลังตรวจสอบเบอร์โทรซ้ำ...'):
                    try:
                        # อ่านไฟล์ Excel
                        df = pd.read_excel(uploaded_file, dtype=str)
                        df = df.rename(columns={df.columns[0]: 'A'})
                        df['A'] = df['A'].astype(str).fillna('')
                        
                        st.info(f"ใช้คอลัมน์แรกเป็นคอลัมน์เบอร์โทร (พบ {len(df)} แถว)")
                        
                        # ดึงตัวเลข 9 ตัวท้าย
                        df['last_9_digits'] = df['A'].apply(extract_last_9_digits)
                        
                        # ดึงข้อมูลเบอร์เก่าจากฐานข้อมูล
                        existing_last_9_digits = get_all_last_9_digits()
                        
                        # ตรวจสอบซ้ำ
                        df['is_duplicate'] = df['last_9_digits'].isin(existing_last_9_digits)
                        
                        # กรองข้อมูลที่ไม่ซ้ำ
                        unique_df = df[~df['is_duplicate']].copy()
                        unique_df = unique_df.drop(columns=['last_9_digits', 'is_duplicate'])
                        
                        # บันทึกลงฐานข้อมูล
                        if save_to_db:
                            new_records = save_phones_to_database(df['A'].tolist(), uploaded_file.name)
                            st.success(f"💾 บันทึกเบอร์โทรลงฐานข้อมูลเรียบร้อย (เพิ่ม {new_records} เบอร์ใหม่)")
                        
                        # แสดงผลลัพธ์
                        st.success("✅ ตรวจสอบเสร็จสิ้น!")
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("เบอร์โทรทั้งหมด", len(df))
                        with col2:
                            st.metric("เบอร์ที่ไม่ซ้ำ", len(unique_df))
                        with col3:
                            st.metric("เบอร์ที่ซ้ำ", len(df) - len(unique_df))
                        
                        # ดาวน์โหลดไฟล์ผลลัพธ์
                        output = save_phones_as_excel(unique_df)
                        
                        original_name = uploaded_file.name
                        name_without_ext = original_name.rsplit('.', 1)[0]
                        download_filename = f"{name_without_ext}-Cut.xlsx"
                        
                        st.download_button(
                            label="💾 ดาวน์โหลดไฟล์ผลลัพธ์",
                            data=output.getvalue(),
                            file_name=download_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                        
                    except Exception as e:
                        st.error(f"❌ เกิดข้อผิดพลาด: {str(e)}")

# เริ่มต้นฐานข้อมูล
if not os.path.exists("phone_database.db"):
    init_database()

# รันแอป
if __name__ == "__main__":
    main()
